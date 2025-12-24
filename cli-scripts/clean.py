#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import logging
import pandas as pd
import numpy as np
import os
from openpyxl.utils import get_column_letter, column_index_from_string

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

def parse_protect_cols(rule):
    """
    支持三种：
    A-F
    A,B,D
    1-6
    """
    rule = rule.replace(" ", "")

    protect = set()

    # 区间 A-F / 1-6
    if "-" in rule:
        start, end = rule.split("-")

        # 字母区间
        if start.isalpha():
            s = column_index_from_string(start)
            e = column_index_from_string(end)
        else:
            s = int(start)
            e = int(end)

        for i in range(s, e + 1):
            protect.add(i)

    # 列表 A,B,C
    elif "," in rule:
        parts = rule.split(",")
        for p in parts:
            if p.isalpha():
                protect.add(column_index_from_string(p))
            else:
                protect.add(int(p))

    # 单值
    else:
        if rule.isalpha():
            protect.add(column_index_from_string(rule))
        else:
            protect.add(int(rule))

    logging.info(f"行清洗保护列: {sorted(protect)}")
    return protect

# region Core Functions
def clean_excel_file(
        input_file,
        sheet_name,
        output_file,
        dirty_file,
        col_threshold,
        row_threshold,
        protect_rule
):
    logging.info(f"加载文件：{input_file}")
    df = pd.read_excel(input_file, sheet_name=sheet_name, dtype=object)
    rows, cols = df.shape
    logging.info(f"表格维度：{rows} 行, {cols} 列")

    dirty_records = [] # 脏记录
    protect_cols = parse_protect_cols(protect_rule) # 行清洗阶段保护列

    # ------------------ 列清洗 ------------------
    logging.info("分析列缺失率...")
    df = df.replace(r'^\s*$', np.nan, regex=True)
    nan_mask = df.isna()
    col_nan_ratio = nan_mask.sum() / len(df)

    cols_to_drop = [col for col, r in col_nan_ratio.items() if r >= col_threshold]
    logging.info(f"符合删除条件的列数量：{len(cols_to_drop)}")

    # DirtyCleanLog 日志记录
    for col in cols_to_drop:
        idx = df.columns.get_loc(col)
        letter = get_column_letter(idx + 1)
        dirty_records.append({
            "type": "column_removed",
            "row": "",
            "col": letter,
            "col_name": col
        })

    df.drop(columns=cols_to_drop, inplace=True)

    # ------------------ 行清洗 ------------------
    logging.info("分析行缺失率...")
    df = df.replace(r'^\s*$', np.nan, regex=True)
    nan_mask = df.isna()
    row_nan_ratio = nan_mask.sum(axis=1) / df.shape[1]

    rows_to_process = row_nan_ratio[row_nan_ratio >= row_threshold].index.tolist()
    logging.info(f"符合清理条件的行数量：{len(rows_to_process)}")

    for row_idx in rows_to_process:
        for col_idx in range(df.shape[1]):
            excel_col_no = col_idx + 1

            # 在保护区，跳过
            if excel_col_no in protect_cols:
                continue

            col_letter = get_column_letter(excel_col_no)

            # DirtyCleanLog 日志记录
            dirty_records.append({
                "type": "row_cell_cleared",
                "row": row_idx + 2,   # +1 数据 +1 header
                "col": col_letter,
                "col_name": df.columns[col_idx]
            })

            df.iat[row_idx, col_idx] = np.nan

    # ------------------ 输出 ------------------
    logging.info("写入清洗结果文件...")
    df.to_excel(output_file, index=False)

    dirty_df = pd.DataFrame(dirty_records)
    if dirty_df.empty:
        dirty_df = pd.DataFrame(columns=["type", "row", "col", "col_name"])
    dirty_df.to_excel(dirty_file, index=False)

    logging.info("任务完成 🎉")
    logging.info(f"清洗输出：{output_file}")
    logging.info(f"异常记录：{dirty_file}")

# ------------------- 目录模式 -------------------
def batch_clean(
        folder, 
        sheet_name, 
        output_dir, 
        dirty_dir,
        col_threshold,
        row_threshold,
        protect_rule
):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    if not os.path.exists(dirty_dir):
        os.makedirs(dirty_dir)

    files = [f for f in os.listdir(folder) if f.endswith(".xlsx") or f.endswith(".xls")]
    if not files:
        logging.warning("目录中没有 Excel 文件！")
        return
    logging.info(f"发现 {len(files)} 个 Excel 文件")

    for f in files:
        input_path = os.path.join(folder, f)
        output_path = os.path.join(output_dir, f)
        dirty_path = os.path.join(dirty_dir, f"dirty_{f}")

        clean_excel_file(
            input_file=input_path,
            sheet_name=sheet_name,
            output_file=output_path,
            dirty_file=dirty_path,
            col_threshold=col_threshold,
            row_threshold=row_threshold,
            protect_rule=protect_rule
        )
# endregion

# ------------------- CLI -------------------
def main():
    parser = argparse.ArgumentParser(description="Excel 清洗工具（支持文件 & 目录）")

    parser.add_argument("-i", "--input", required=True, help="输入文件 或 文件夹")
    parser.add_argument("-s", "--sheet", required=True, help="Sheet 名称")
    parser.add_argument("-o", "--output", required=False, help="输出文件 或 输出文件夹")
    parser.add_argument("--dirty", required=False, help="异常输出目录")

    parser.add_argument("--col-threshold", type=float, default=0.7, help="列删除阈值(默认 0.7)")
    parser.add_argument("--row-threshold", type=float, default=0.7, help="行清理阈值(默认 0.7)")
    parser.add_argument("--protect-cols", default="A-F", help="行清洗保留列 (A-F / 1-6 / A,B,D)")

    args = parser.parse_args()

    # ------------ 自动生成默认输出路径 ------------
    input_path = args.input
    if os.path.isdir(input_path):
        base_name = os.path.basename(os.path.normpath(input_path))
        if not args.output:
            args.output = os.path.join(
                os.path.dirname(input_path),
                f"clean_{base_name}"
            )
        if not args.dirty or args.dirty == "dirty_output":
            args.dirty = os.path.join(
                os.path.dirname(input_path),
                f"dirty_{base_name}"
            )
    else:
        # 输入是文件
        dir_name = os.path.dirname(input_path)
        file_name = os.path.basename(input_path)

        if not args.output:
            args.output = os.path.join(dir_name, f"clean_{file_name}")

        if not args.dirty or args.dirty == "dirty_output":
            args.dirty = os.path.join(dir_name, f"dirty_{file_name}")

    # ------------ 处理逻辑分支 ------------
    if os.path.isdir(args.input):
        logging.info("进入批量处理模式")
        batch_clean(
            folder=args.input,
            sheet_name=args.sheet,
            output_dir=args.output,
            dirty_dir=args.dirty,
            col_threshold=args.col_threshold,
            row_threshold=args.row_threshold,
            protect_rule=args.protect_cols
        )
    else:
        logging.info("进入单文件模式")
        clean_excel_file(
            input_file=args.input,
            sheet_name=args.sheet,
            output_file=args.output,
            dirty_file=args.dirty,
            col_threshold=args.col_threshold,
            row_threshold=args.row_threshold,
            protect_rule=args.protect_cols
        )


if __name__ == "__main__":
    main()
